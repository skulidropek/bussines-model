#!/usr/bin/env python3
"""
Independent FCFF/FCFE recomputation from mapped workbook rows.

Usage:
  python3 recalc_fcff_fcfe.py --workbook model.xlsx --config mappings.json --output metric_reconciliation.json
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def get_num(v: Any) -> float:
    if v in (None, ""):
        return 0.0
    return float(v)


def resolve_sheet(wb_formula, wb_values, requested: str | None):
    if requested and requested in wb_formula.sheetnames:
        return wb_formula[requested], wb_values[requested], requested, False
    fallback = wb_formula.sheetnames[0]
    return wb_formula[fallback], wb_values[fallback], fallback, bool(requested)


def npv(series: list[float], rates: list[float], years: list[int], base_year: int, shift: int) -> float:
    out = 0.0
    for i, cf in enumerate(series):
        t = (years[i] - base_year) + shift
        out += cf / ((1 + rates[i]) ** t)
    return out


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--config", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    cfg = json.loads(Path(args.config).read_text(encoding="utf-8"))
    wb_formula = load_workbook(args.workbook, data_only=False)
    wb_values = load_workbook(args.workbook, data_only=True)
    requested_sheet = cfg.get("sheet")
    ws, ws_values, _, used_fallback = resolve_sheet(wb_formula, wb_values, requested_sheet)

    cols = cfg["columns"]
    year_row = int(cfg["year_row"])
    rows = cfg["rows"]

    def num_at(addr: str) -> float:
        v_formula = ws[addr].value
        if isinstance(v_formula, str) and v_formula.startswith("="):
            v_val = ws_values[addr].value
            if v_val in (None, ""):
                raise ValueError(
                    f"Cell {addr} has formula but no cached numeric value. "
                    "Open workbook in Excel/Sheets and recalculate before running this script."
                )
            return get_num(v_val)
        return get_num(v_formula)

    years = [int(num_at(f"{c}{year_row}")) for c in cols]
    base_year = int(cfg.get("base_year", years[0]))
    discount_shift = int(cfg.get("discount_shift", 1))

    ebit = [num_at(f"{c}{rows['ebit']}") for c in cols]
    tax = [num_at(f"{c}{rows['tax']}") for c in cols]
    d_ar = [num_at(f"{c}{rows['delta_ar']}") for c in cols]
    d_ap = [num_at(f"{c}{rows['delta_ap']}") for c in cols]
    capex = [num_at(f"{c}{rows['capex']}") for c in cols]
    interest = [num_at(f"{c}{rows['interest']}") for c in cols]
    draw = [num_at(f"{c}{rows['debt_draw']}") for c in cols]
    repay = [num_at(f"{c}{rows['debt_repay']}") for c in cols]
    divs = [num_at(f"{c}{rows['dividends']}") for c in cols]
    eq_in = [num_at(f"{c}{rows['equity_in']}") for c in cols]
    rates = [num_at(f"{c}{rows['discount_rate']}") for c in cols]

    fcff = []
    fcfe_dist = []
    fcfe_formula = []

    interest_after_tax = bool(cfg.get("interest_after_tax", False))
    tax_rate = float(cfg.get("tax_rate", 0.0))

    for i in range(len(cols)):
        cfo_unlevered = ebit[i] - tax[i] - d_ar[i] + d_ap[i]
        fcff_i = cfo_unlevered - capex[i]
        net_borrow = draw[i] - repay[i]

        if interest_after_tax:
            int_term = interest[i] * (1 - tax_rate)
        else:
            int_term = interest[i]

        fcfe_i = fcff_i - int_term + net_borrow
        fcff.append(fcff_i)
        fcfe_dist.append(-eq_in[i] + divs[i])
        fcfe_formula.append(fcfe_i)

    npv_fcff = npv(fcff, rates, years, base_year, discount_shift)
    npv_fcfe_dist = npv(fcfe_dist, rates, years, base_year, discount_shift)
    npv_fcfe_formula = npv(fcfe_formula, rates, years, base_year, discount_shift)

    def model_value(cell_name: str | None) -> float | None:
        if not cell_name:
            return None
        return num_at(cell_name)

    model_npv_fcff = model_value(cfg.get("model_npv_fcff_cell"))
    model_npv_fcfe = model_value(cfg.get("model_npv_fcfe_cell"))

    tol = float(cfg.get("tolerance_pct", 0.1)) / 100.0

    def pct_diff(a: float, b: float | None) -> float | None:
        if b is None:
            return None
        base = abs(b) if abs(b) > 1e-9 else 1.0
        return (a - b) / base

    diff_fcff = pct_diff(npv_fcff, model_npv_fcff)
    diff_fcfe = pct_diff(npv_fcfe_dist, model_npv_fcfe)

    pass_fcff = diff_fcff is None or abs(diff_fcff) <= tol
    pass_fcfe = diff_fcfe is None or abs(diff_fcfe) <= tol

    payload = {
        "workbook": args.workbook,
        "sheet": ws.title,
        "requested_sheet": requested_sheet,
        "used_sheet_fallback": used_fallback,
        "available_sheets": wb_formula.sheetnames,
        "years": years,
        "assumptions": {
            "base_year": base_year,
            "discount_shift": discount_shift,
            "interest_after_tax": interest_after_tax,
            "tax_rate": tax_rate,
            "tolerance_pct": tol * 100,
        },
        "series": {
            "fcff": fcff,
            "fcfe_distribution": fcfe_dist,
            "fcfe_formula": fcfe_formula,
        },
        "npv": {
            "recalc_fcff": npv_fcff,
            "recalc_fcfe_distribution": npv_fcfe_dist,
            "recalc_fcfe_formula": npv_fcfe_formula,
            "model_fcff": model_npv_fcff,
            "model_fcfe": model_npv_fcfe,
            "diff_fcff_pct": diff_fcff,
            "diff_fcfe_pct": diff_fcfe,
        },
        "pass": {
            "fcff": pass_fcff,
            "fcfe": pass_fcfe,
            "overall": pass_fcff and pass_fcfe,
        },
    }

    out = Path(args.output)
    out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {out}; pass={payload['pass']['overall']}")

    if not payload["pass"]["overall"]:
        raise SystemExit(3)


if __name__ == "__main__":
    main()
