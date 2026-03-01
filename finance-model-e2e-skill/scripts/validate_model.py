#!/usr/bin/env python3
"""
Validate workbook quality gates and explicit checks from a JSON config.

Usage:
  python3 validate_model.py --workbook model.xlsx --config validation_config.json --output validation_report.json
"""

from __future__ import annotations

import argparse
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


@dataclass
class CheckResult:
    name: str
    passed: bool
    detail: str


def get_sheet(wb, name: str | None):
    if name and name in wb.sheetnames:
        return wb[name], False
    return wb[wb.sheetnames[0]], bool(name)


def scan_formula_errors(ws) -> list[str]:
    bad = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith("=") and "#REF!" in v.upper():
                bad.append(cell.coordinate)
    return bad


def run_check(ws_formula, ws_values, check: dict[str, Any]) -> CheckResult:
    ctype = check["type"]
    name = check.get("name", ctype)

    def numeric_at(addr: str) -> float:
        v_formula = ws_formula[addr].value
        if isinstance(v_formula, str) and v_formula.startswith("="):
            v_val = ws_values[addr].value
            if v_val in (None, ""):
                raise ValueError(
                    f"Cell {addr} has formula but no cached numeric value. "
                    "Recalculate workbook in Excel/Sheets before validation."
                )
            return float(v_val)
        v = ws_formula[addr].value
        return 0.0 if v is None else float(v)

    if ctype == "equals_zero":
        tol = float(check.get("tolerance", 0.0))
        failures = []
        for addr in check["cells"]:
            val = numeric_at(addr)
            if abs(val) > tol:
                failures.append((addr, val))
        passed = not failures
        detail = "OK" if passed else f"Non-zero cells: {failures}"
        return CheckResult(name, passed, detail)

    if ctype == "text_equals":
        addr = check["cell"]
        expected = str(check["expected"])
        actual = str(ws_formula[addr].value)
        passed = actual == expected
        return CheckResult(name, passed, f"{addr}: actual={actual!r} expected={expected!r}")

    if ctype == "formula_contains":
        addr = check["cell"]
        needle = str(check["contains"])
        formula = ws_formula[addr].value
        passed = isinstance(formula, str) and formula.startswith("=") and needle in formula
        return CheckResult(name, passed, f"{addr}: formula={formula!r} needle={needle!r}")

    if ctype == "not_blank":
        missing = [addr for addr in check["cells"] if ws_formula[addr].value in (None, "")]
        passed = not missing
        return CheckResult(name, passed, "OK" if passed else f"Blank cells: {missing}")

    return CheckResult(name, False, f"Unsupported check type: {ctype}")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--config", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    cfg = json.loads(Path(args.config).read_text(encoding="utf-8"))
    wb_formula = load_workbook(args.workbook, data_only=False)
    wb_values = load_workbook(args.workbook, data_only=True)
    requested_sheet = cfg.get("sheet")
    ws, used_fallback = get_sheet(wb_formula, requested_sheet)
    ws_values, _ = get_sheet(wb_values, requested_sheet)

    results: list[CheckResult] = []

    # Global formula integrity scan
    ref_errors = scan_formula_errors(ws)
    results.append(
        CheckResult(
            name="formula_ref_errors",
            passed=not ref_errors,
            detail="OK" if not ref_errors else f"#REF! found in cells: {ref_errors}",
        )
    )

    for check in cfg.get("checks", []):
        results.append(run_check(ws, ws_values, check))

    passed = all(r.passed for r in results)
    payload = {
        "workbook": args.workbook,
        "sheet": ws.title,
        "requested_sheet": requested_sheet,
        "used_sheet_fallback": used_fallback,
        "available_sheets": wb_formula.sheetnames,
        "passed": passed,
        "checks": [r.__dict__ for r in results],
    }

    out = Path(args.output)
    out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {out}; passed={passed}")

    if not passed:
        raise SystemExit(2)


if __name__ == "__main__":
    main()
